### Import required modules

import requests
import os.path
import csv
import openpyxl


### Download the .csv file from the web

#url = 'https://www.stats.govt.nz/assets/Uploads/Annual-enterprise-survey/Annual-enterprise-survey-2019-financial-year-provisional/Download-data/annual-enterprise-survey-2019-financial-year-provisional-size-bands-csv.csv'
#r = requests.get(url, allow_redirects=True)

#user_path = 'C:/Users/' + input('Enter Windows username: ') + '/'
#user_path = 'C:/Users/' + '970jwillems' + '/'
#folder = 'Sonova/Sonova Marketing GmbH - Data and Analytics/_Janine'
#filename =  str((url.rsplit("/",1)[1:][0]))
    #get last part of url after / with .rsplit("/",1)
    #then from this list containing the first item [1:] of the result of .rsplit() 
    #return the first element with [0] as a string

#save_path = os.path.join(user_path, folder, filename)
#open(save_path,'wb+').write(r.content)
    #w for write, b for binary file, + for creating file if not exists


### Convert .csv to .xlsx file

wb = openpyxl.Workbook()
ws = wb.active
location = 'C:\\Users\\970jwillems\\OneDrive - Sonova\\Reporting-Dashboards\\CCC\\DASD-71 Forecast CCC\\'
csvfile =  'ccc_forecast_202010091219.csv'
csvfilepath = os.path.join(location, csvfile)
xlfilepath = os.path.join(location, csvfile) [:-4] + '.xlsx'


reader = csv.reader(csvfile, delimiter=',')
sortedfile = sorted(reader, key=lambda column: column[1], reverse=True)

with open(csvfilepath) as csvfile:
    for row in sortedfile:
        ws.append(row)

wb.save(xlfilepath)
wb.close()


### Customise the .xlsx file

open(xlfilename)

ws.title = "Default Scenario"

wb.copy_worksheet(wb.worksheets[0])
wb.worksheets[1].title = "Scenario Planning"
    #copy the worksheet and change the name

wb.create_sheet("Scenario Planning Parameters")

for row in wb.worksheets[1]:
    for cell in row:
        wb.worksheets[2][cell.coordinate].value = cell.value
            #fill a worksheet (or any range from anywhere - more flexible method!)
            #by grasping the cell value from another worksheet
        
parsheet = wb.worksheets[2]

for i, cellObj in enumerate(parsheet['F2:F' + str(parsheet.max_row)],1):
    try:
        int(cellObj[0].value)
            #enumerate(iterable, start index)
            #cellObj is in this case also a tuple (why??), return the first element of this tuple
            #access the value
        
    except ValueError:
        pass

for i, cellObj in enumerate(parsheet['I2:I' + str(parsheet.max_row)], 2):
    cellObj[0].value = '=$F${0} + 2'.format(i)

for cellObj in parsheet['C']:
    cellObj.value = None 

wb.save(xlfilename)
wb.close()
